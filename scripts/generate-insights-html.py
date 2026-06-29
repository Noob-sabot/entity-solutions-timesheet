#!/usr/bin/env python3
"""Generate a self-contained insights-repository.html from the Excel workbook."""

from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path

import openpyxl

DEFAULT_SRC = Path("/Users/transport/Downloads/1.0 Insights Repository.backup.xlsx")
DEFAULT_OUT = Path(__file__).resolve().parents[1] / "insights-repository.html"

KEY_MAP = {
    "# Insight": "id",
    "Trip Motivation": "tripMotivation",
    "Transport Mode": "transportMode",
    "Context": "context",
    "Plan": "plan",
    "Travel to stop": "travelToStop",
    "Arrive at stop": "arriveAtStop",
    "Wait for service": "waitForService",
    "Travel on-board": "travelOnBoard",
    "Disembark and interchange": "disembark",
    "Post Trip": "postTrip",
    "Insight source": "source",
    "Ref.": "ref",
    "Type": "type",
    "Primary Theme": "primaryTheme",
    "Secondary Theme": "secondaryTheme",
    "Conceptual Theme": "conceptualTheme",
    "Insight": "insight",
}

JOURNEY_STAGES = [
    ("plan", "Plan"),
    ("travelToStop", "Travel to stop"),
    ("arriveAtStop", "Arrive at stop"),
    ("waitForService", "Wait for service"),
    ("travelOnBoard", "Travel on-board"),
    ("disembark", "Disembark and interchange"),
    ("postTrip", "Post Trip"),
]

FILTER_FIELDS = [
    ("transportMode", "Transport Mode"),
    ("tripMotivation", "Trip Motivation"),
    ("context", "Context"),
    ("type", "Type"),
    ("primaryTheme", "Primary Theme"),
    ("secondaryTheme", "Secondary Theme"),
    ("conceptualTheme", "Conceptual Theme"),
    ("source", "Insight source"),
]

HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Insights Repository</title>
  <style>
    :root {
      --bg: #f4f6f8;
      --surface: #fff;
      --border: #d8dee6;
      --text: #1a2332;
      --muted: #5c6b7a;
      --accent: #0065bd;
      --accent-soft: #e8f2fb;
      --header: #0f2940;
      --ok: #1a7f4b;
      --shadow: 0 1px 3px rgba(15, 41, 64, 0.08);
    }
    * { box-sizing: border-box; }
    body {
      margin: 0;
      font-family: "Segoe UI", system-ui, -apple-system, sans-serif;
      background: var(--bg);
      color: var(--text);
      line-height: 1.45;
    }
    header {
      background: var(--header);
      color: #fff;
      padding: 1rem 1.5rem 0;
      box-shadow: var(--shadow);
    }
    header h1 { margin: 0 0 0.75rem; font-size: 1.35rem; font-weight: 600; }
    nav { display: flex; gap: 0.25rem; flex-wrap: wrap; }
    nav button {
      background: transparent;
      border: none;
      color: rgba(255,255,255,0.75);
      padding: 0.65rem 1rem;
      cursor: pointer;
      border-radius: 6px 6px 0 0;
      font: inherit;
      font-size: 0.95rem;
    }
    nav button.active {
      background: var(--bg);
      color: var(--header);
      font-weight: 600;
    }
    main { max-width: 1600px; margin: 0 auto; padding: 1.25rem 1.5rem 2rem; }
    .panel {
      background: var(--surface);
      border: 1px solid var(--border);
      border-radius: 10px;
      box-shadow: var(--shadow);
      padding: 1rem 1.25rem;
      margin-bottom: 1rem;
    }
    .panel h2 { margin: 0 0 0.5rem; font-size: 1.05rem; }
    .panel p { margin: 0.35rem 0; color: var(--muted); }
    .stats { display: flex; gap: 1.5rem; flex-wrap: wrap; margin-top: 0.75rem; }
    .stat strong { display: block; font-size: 1.4rem; color: var(--accent); }
    .stat span { font-size: 0.85rem; color: var(--muted); }
    .layout { display: grid; grid-template-columns: 280px 1fr; gap: 1rem; align-items: start; }
    @media (max-width: 960px) { .layout { grid-template-columns: 1fr; } }
    .filters { position: sticky; top: 1rem; max-height: calc(100vh - 2rem); overflow: auto; }
    .filter-group { margin-bottom: 1rem; }
    .filter-group label.title {
      display: block;
      font-size: 0.78rem;
      font-weight: 700;
      text-transform: uppercase;
      letter-spacing: 0.04em;
      color: var(--muted);
      margin-bottom: 0.35rem;
    }
    .filter-group select {
      width: 100%;
      min-height: 88px;
      border: 1px solid var(--border);
      border-radius: 6px;
      padding: 0.35rem;
      font: inherit;
      font-size: 0.88rem;
    }
    .stage-grid { display: grid; gap: 0.25rem; }
    .stage-grid label {
      display: flex;
      align-items: center;
      gap: 0.4rem;
      font-size: 0.88rem;
      cursor: pointer;
    }
    .toolbar {
      display: flex;
      gap: 0.5rem;
      flex-wrap: wrap;
      align-items: center;
      margin-bottom: 0.75rem;
    }
    input[type="search"] {
      flex: 1;
      min-width: 200px;
      border: 1px solid var(--border);
      border-radius: 6px;
      padding: 0.5rem 0.75rem;
      font: inherit;
    }
    button.action {
      border: 1px solid var(--border);
      background: #fff;
      border-radius: 6px;
      padding: 0.45rem 0.85rem;
      font: inherit;
      cursor: pointer;
    }
    button.action.primary { background: var(--accent); color: #fff; border-color: var(--accent); }
    button.action:hover { filter: brightness(0.97); }
    .preset-row { display: flex; flex-wrap: wrap; gap: 0.35rem; margin-bottom: 0.75rem; }
    .preset-row button { font-size: 0.82rem; padding: 0.3rem 0.6rem; }
    .table-wrap { overflow: auto; border: 1px solid var(--border); border-radius: 8px; max-height: 70vh; }
    table { width: 100%; border-collapse: collapse; font-size: 0.84rem; }
    th, td { padding: 0.45rem 0.55rem; border-bottom: 1px solid var(--border); vertical-align: top; text-align: left; }
    th {
      position: sticky;
      top: 0;
      background: #eef3f7;
      z-index: 1;
      font-size: 0.78rem;
      text-transform: uppercase;
      letter-spacing: 0.03em;
      white-space: nowrap;
    }
    tr:hover td { background: #fafcfd; }
    td.insight { min-width: 280px; max-width: 420px; }
    td.check { text-align: center; color: var(--ok); }
    .hidden { display: none !important; }
    a { color: var(--accent); }
    .notice {
      background: var(--accent-soft);
      border: 1px solid #b9d9f5;
      border-radius: 8px;
      padding: 0.75rem 1rem;
      margin-bottom: 1rem;
      font-size: 0.92rem;
    }
    .empty { padding: 2rem; text-align: center; color: var(--muted); }
  </style>
</head>
<body>
  <header>
    <h1>Insights Repository</h1>
    <nav id="tabs"></nav>
  </header>
  <main id="app"></main>
  <script>
    const DATA = __DATA__;

    const TABS = [
      { id: 'overview', label: 'Overview' },
      { id: 'insights', label: 'Insights' },
      { id: 'sources', label: 'Sources' },
      { id: 'filtered-sources', label: 'Filtered Sources' },
    ];

    const state = {
      tab: 'insights',
      search: '',
      filters: Object.fromEntries(DATA.filterFields.map(([k]) => [k, new Set()])),
      stages: new Set(),
    };

    function $(sel, root = document) { return root.querySelector(sel); }
    function el(tag, props = {}, children = []) {
      const node = document.createElement(tag);
      Object.entries(props).forEach(([k, v]) => {
        if (k === 'className') node.className = v;
        else if (k === 'text') node.textContent = v;
        else if (k.startsWith('on')) node.addEventListener(k.slice(2).toLowerCase(), v);
        else node.setAttribute(k, v);
      });
      children.forEach(c => node.append(c instanceof Node ? c : document.createTextNode(c)));
      return node;
    }

    function hasCheck(v) { return v === '✓' || v === true || v === 'x' || v === 'X'; }

    function rowMatches(row) {
      for (const [key] of DATA.filterFields) {
        const selected = state.filters[key];
        if (selected.size && !selected.has(row[key])) return false;
      }
      for (const stage of state.stages) {
        if (!hasCheck(row[stage])) return false;
      }
      if (state.search) {
        const q = state.search.toLowerCase();
        const hay = [row.insight, row.id, row.source, row.primaryTheme, row.secondaryTheme]
          .filter(Boolean).join(' ').toLowerCase();
        if (!hay.includes(q)) return false;
      }
      return true;
    }

    function filteredInsights() { return DATA.insights.filter(rowMatches); }

    function filteredSources() {
      const ids = [...new Set(filteredInsights().map(r => r.source).filter(Boolean))];
      const byId = new Map(DATA.sources.map(s => [s.id, s]));
      return ids.map(id => byId.get(id)).filter(Boolean);
    }

    function renderTabs() {
      const nav = $('#tabs');
      nav.replaceChildren(...TABS.map(t => el('button', {
        className: state.tab === t.id ? 'active' : '',
        text: t.label,
        onClick: () => { state.tab = t.id; render(); },
      })));
    }

    function renderOverview() {
      return el('div', {}, [
        el('div', { className: 'panel' }, [
          el('h2', { text: 'About this tool' }),
          el('p', { text: 'This page replaces the Excel Insights Repository for browsing and filtering customer insights and finding the source documents you need to review.' }),
        ]),
        el('div', { className: 'panel' }, [
          el('h2', { text: 'Insights' }),
          el('p', { text: 'Browse all customer insights. Use the filters on the left to narrow by transport mode, trip motivation, theme, journey stage, and more. Journey stage filters show insights marked with a check for that stage.' }),
        ]),
        el('div', { className: 'panel' }, [
          el('h2', { text: 'Sources' }),
          el('p', { text: 'Full reference list of all source documents — ID, title, author, date, and intranet link where available.' }),
        ]),
        el('div', { className: 'panel' }, [
          el('h2', { text: 'Filtered Sources' }),
          el('p', { text: 'Automatically lists source documents referenced by insights that match your current filters. Change filters on the Insights tab (or here — filters are shared) and the document list updates instantly. Each source appears once, with the same columns as Sources.' }),
        ]),
      ]);
    }

    function filterSelect(key, label) {
      const select = el('select', { multiple: 'multiple', size: '5' });
      DATA.options[key].forEach(opt => {
        const o = el('option', { value: opt, text: opt });
        if (state.filters[key].has(opt)) o.selected = true;
        select.append(o);
      });
      select.addEventListener('change', () => {
        state.filters[key] = new Set([...select.selectedOptions].map(o => o.value));
        render();
      });
      return el('div', { className: 'filter-group' }, [
        el('label', { className: 'title', text: label }),
        select,
      ]);
    }

    function renderFilters() {
      const wrap = el('div', { className: 'panel filters' }, [
        el('h2', { text: 'Filters' }),
        el('p', { text: 'Select one or more values per field. Filters combine with AND logic.' }),
      ]);
      DATA.filterFields.forEach(([key, label]) => wrap.append(filterSelect(key, label)));

      const stageGrid = el('div', { className: 'stage-grid' });
      DATA.journeyStages.forEach(([key, label]) => {
        const cb = el('input', { type: 'checkbox' });
        cb.checked = state.stages.has(key);
        cb.addEventListener('change', () => {
          if (cb.checked) state.stages.add(key); else state.stages.delete(key);
          render();
        });
        stageGrid.append(el('label', {}, [cb, label]));
      });
      wrap.append(el('div', { className: 'filter-group' }, [
        el('label', { className: 'title', text: 'Journey stage (has checkmark)' }),
        stageGrid,
      ]));

      wrap.append(el('button', {
        className: 'action',
        text: 'Clear all filters',
        onClick: () => {
          Object.keys(state.filters).forEach(k => state.filters[k] = new Set());
          state.stages.clear();
          state.search = '';
          render();
        },
      }));
      return wrap;
    }

    function applyPresetBusCoach() {
      state.filters.transportMode = new Set([
        'Metro and regional town bus', 'Metro bus', 'Regional town bus',
        'V/Line coach', 'Multiple', 'V/Line (coach and train)',
      ]);
      render();
    }

    function renderInsightsTable(rows) {
      const cols = [
        ['id', '# Insight'], ['tripMotivation', 'Trip Motivation'], ['transportMode', 'Transport Mode'],
        ['context', 'Context'], ...DATA.journeyStages, ['source', 'Insight source'], ['ref', 'Ref.'],
        ['type', 'Type'], ['primaryTheme', 'Primary Theme'], ['secondaryTheme', 'Secondary Theme'],
        ['conceptualTheme', 'Conceptual Theme'], ['insight', 'Insight'],
      ];
      const thead = el('tr', {}, cols.map(([, label]) => el('th', { text: label })));
      const tbody = el('tbody');
      if (!rows.length) {
        tbody.append(el('tr', {}, [el('td', { colSpan: String(cols.length), className: 'empty', text: 'No insights match the current filters.' })]));
      } else {
        rows.forEach(row => {
          tbody.append(el('tr', {}, cols.map(([key]) => {
            const val = row[key];
            const isStage = DATA.journeyStages.some(([k]) => k === key);
            const td = el('td', {
              className: key === 'insight' ? 'insight' : (isStage ? 'check' : ''),
              text: isStage ? (hasCheck(val) ? '✓' : '') : (val ?? ''),
            });
            return td;
          })));
        });
      }
      return el('div', { className: 'table-wrap' }, [el('table', {}, [el('thead', {}, [thead]), tbody])]);
    }

    function renderInsights() {
      const rows = filteredInsights();
      const content = el('div', {}, [
        el('div', { className: 'panel' }, [
          el('div', { className: 'stats' }, [
            el('div', { className: 'stat' }, [el('strong', { text: String(rows.length) }), el('span', { text: 'Matching insights' })]),
            el('div', { className: 'stat' }, [el('strong', { text: String(filteredSources().length) }), el('span', { text: 'Source documents' })]),
          ]),
        ]),
        el('div', { className: 'preset-row' }, [
          el('span', { text: 'Quick filter: ', style: 'font-size:0.85rem;color:var(--muted);align-self:center;' }),
          el('button', { className: 'action', text: 'Bus & coach (+ Multiple, V/Line both)', onClick: applyPresetBusCoach }),
        ]),
        el('div', { className: 'toolbar' }, [
          el('input', {
            type: 'search',
            placeholder: 'Search insight text, ID, source, theme…',
            value: state.search,
            onInput: e => { state.search = e.target.value; renderInsightsMain(); },
          }),
          el('button', { className: 'action', text: 'View filtered sources', onClick: () => { state.tab = 'filtered-sources'; render(); } }),
        ]),
        renderInsightsTable(rows),
      ]);
      return el('div', { className: 'layout' }, [renderFilters(), content]);
    }

    function renderInsightsMain() {
      const mount = $('#insights-main');
      if (!mount) return;
      const rows = filteredInsights();
      mount.replaceChildren(
        el('div', { className: 'panel' }, [
          el('div', { className: 'stats' }, [
            el('div', { className: 'stat' }, [el('strong', { text: String(rows.length) }), el('span', { text: 'Matching insights' })]),
            el('div', { className: 'stat' }, [el('strong', { text: String(filteredSources().length) }), el('span', { text: 'Source documents' })]),
          ]),
        ]),
        el('div', { className: 'preset-row' }, [
          el('span', { text: 'Quick filter: ', style: 'font-size:0.85rem;color:var(--muted);align-self:center;' }),
          el('button', { className: 'action', text: 'Bus & coach (+ Multiple, V/Line both)', onClick: applyPresetBusCoach }),
        ]),
        el('div', { className: 'toolbar' }, [
          el('input', {
            type: 'search',
            placeholder: 'Search insight text, ID, source, theme…',
            value: state.search,
            onInput: e => { state.search = e.target.value; renderInsightsMain(); },
          }),
          el('button', { className: 'action', text: 'View filtered sources', onClick: () => { state.tab = 'filtered-sources'; render(); } }),
        ]),
        renderInsightsTable(rows),
      );
    }

    function renderSourcesTable(rows, notice) {
      const parts = [];
      if (notice) parts.push(el('div', { className: 'notice', text: notice }));
      parts.push(el('div', { className: 'panel' }, [
        el('div', { className: 'stats' }, [
          el('div', { className: 'stat' }, [el('strong', { text: String(rows.length) }), el('span', { text: 'Documents' })]),
          el('div', { className: 'stat' }, [el('strong', { text: String(filteredInsights().length) }), el('span', { text: 'Matching insights' })]),
        ]),
      ]));
      const thead = el('tr', {}, ['ID#', 'Document Title', 'Author / Org', 'Date published', 'Intranet Link']
        .map(h => el('th', { text: h })));
      const tbody = el('tbody');
      if (!rows.length) {
        tbody.append(el('tr', {}, [el('td', { colSpan: '5', className: 'empty', text: 'No matching sources for the current filters.' })]));
      } else {
        rows.forEach(s => {
          const linkCell = s.link
            ? el('a', { href: s.link, text: s.link, target: '_blank', rel: 'noopener' })
            : document.createTextNode('');
          tbody.append(el('tr', {}, [
            el('td', { text: s.id ?? '' }),
            el('td', { text: s.title ?? '' }),
            el('td', { text: s.author ?? '' }),
            el('td', { text: s.date ?? '' }),
            el('td', {}, [linkCell]),
          ]));
        });
      }
      parts.push(el('div', { className: 'table-wrap' }, [el('table', {}, [el('thead', {}, [thead]), tbody])]));
      return el('div', {}, parts);
    }

    function renderSources() {
      return renderSourcesTable(DATA.sources, null);
    }

    function renderFilteredSources() {
      const notice = 'These documents are based on your current filters (shared with the Insights tab). Each source ID comes from the Insight source column of matching insights. Change filters to explore a different business question.';
      return el('div', { className: 'layout' }, [
        renderFilters(),
        renderSourcesTable(filteredSources(), notice),
      ]);
    }

    function render() {
      renderTabs();
      const app = $('#app');
      app.replaceChildren();
      if (state.tab === 'overview') app.append(renderOverview());
      else if (state.tab === 'insights') {
        const layout = renderInsights();
        layout.querySelector('div.panel + .preset-row')?.previousElementSibling;
        const main = layout.children[1];
        main.id = 'insights-main';
        app.append(layout);
      }
      else if (state.tab === 'sources') app.append(renderSources());
      else if (state.tab === 'filtered-sources') app.append(renderFilteredSources());
    }

    render();
  </script>
</body>
</html>
"""


def normalize(value):
    if isinstance(value, (datetime, date)):
        if isinstance(value, datetime) and value.hour == 0 and value.minute == 0:
            return value.strftime("%b %Y") if value.day == 1 else value.date().isoformat()
        return value.isoformat()
    if value is None or value == "":
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        return value.strip()
    return value


def load_payload(src: Path) -> dict:
    wb = openpyxl.load_workbook(src, data_only=True)

    ws = wb["Insights"]
    headers = [ws.cell(2, c).value for c in range(1, 20)]
    insights = []
    for r in range(3, ws.max_row + 1):
        row = {}
        for c, header in enumerate(headers, 1):
            if header not in KEY_MAP:
                continue
            row[KEY_MAP[header]] = normalize(ws.cell(r, c).value)
        if row.get("id"):
            insights.append(row)

    ws = wb["Sources"]
    sources = []
    for r in range(2, ws.max_row + 1):
        sources.append(
            {
                "id": normalize(ws.cell(r, 1).value),
                "title": normalize(ws.cell(r, 2).value),
                "author": normalize(ws.cell(r, 3).value),
                "date": normalize(ws.cell(r, 4).value),
                "link": normalize(ws.cell(r, 5).value),
            }
        )

    options = {
        key: sorted({row[key] for row in insights if row.get(key)}) for key, _ in FILTER_FIELDS
    }

    return {
        "insights": insights,
        "sources": sources,
        "journeyStages": JOURNEY_STAGES,
        "filterFields": FILTER_FIELDS,
        "options": options,
    }


def generate(src: Path = DEFAULT_SRC, out: Path = DEFAULT_OUT) -> Path:
    if not src.exists():
        raise FileNotFoundError(f"Workbook not found: {src}")
    payload = load_payload(src)
    data_js = json.dumps(payload, ensure_ascii=False)
    html = HTML_TEMPLATE.replace("__DATA__", data_js)
    out.write_text(html, encoding="utf-8")
    return out


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--src", type=Path, default=DEFAULT_SRC)
    parser.add_argument("--out", type=Path, default=DEFAULT_OUT)
    args = parser.parse_args()
    path = generate(args.src, args.out)
    print(f"Wrote {path} ({path.stat().st_size // 1024} KB)")
